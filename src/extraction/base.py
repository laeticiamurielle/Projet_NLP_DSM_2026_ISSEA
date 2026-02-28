"""
audit_snd30.extraction.base
===========================
Fonctions utilitaires partagées par lf_2024.py et lf_2025.py.
"""

import re
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Corrections OCR communes ──────────────────────────────────────────────────
OCR_CORRECTIONS: list[tuple[str, str]] = [
    ("COwORDINATION",      "COORDINATION"),
    ("MULTJLATERALE",      "MULTILATÉRALE"),
    ("D.ECENTRALJSEE",     "DÉCENTRALISÉE"),
    ("D.ECENTRALISEE",     "DÉCENTRALISÉE"),
    ("GOUVERNENMENTALE",   "GOUVERNEMENTALE"),
    ("GOUVERNEMENTA LE",   "GOUVERNEMENTALE"),
    ("NOWELLES",           "NOUVELLES"),
    ("AMELIORATION D",     "AMÉLIORATION D"),
    ("INSTITUTIONELLES",   "INSTITUTIONNELLES"),
    ("AUXw ",              "AUX "),
    ("CONSEI ",            "CONSEIL "),
    ("INTEG RITE",         "INTÉGRITÉ"),
    ("INTEG ",             "INTÉGRITÉ "),
    ("DEL' ",              "DE L'"),
    ("SND30.",             ""),
    (" SPM",               ""),
    (" MINREX",            ""),
]


def clean_ocr(text: str) -> str:
    """Applique les corrections OCR et normalise les espaces."""
    for old, new in OCR_CORRECTIONS:
        text = text.replace(old, new)
    text = re.sub(r"\bw([A-ZÀÂÉÈÊÙÛÔÎŒÆ])", r"\1", text)
    text = re.sub(r"(?<!\w)[mcwrp](?!\w)", "", text)
    return re.sub(r"\s+", " ", text).strip()


def clean_lib(text: str) -> str:
    """Nettoie un libellé (artefacts OCR, espaces multiples)."""
    text = re.sub(r"\b[wW]\b", " ", text)
    text = re.sub(r"\.{3,}", " ", text)
    return re.sub(r"\s{2,}", " ", text).strip()


def parse_amount(word_texts: list[str]) -> int | None:
    """Reconstitue un entier depuis des fragments de texte OCR."""
    raw = "".join(word_texts)
    raw = re.sub(r"\(X+\)", "000", raw)
    raw = re.sub(r"\([0O]{2,3}\)", "000", raw)
    digits = re.sub(r"[^\d]", "", raw)
    if not digits:
        return None
    try:
        return int(digits)
    except ValueError:
        return None


def digs(s: str) -> str:
    """Retourne uniquement les chiffres d'une chaîne."""
    return re.sub(r"[^\d]", "", s)


# ── Export Excel formaté ──────────────────────────────────────────────────────
def export_excel_formate(df: pd.DataFrame, path: Path, sheet_name: str = "Lignes Budgétaires") -> None:
    """
    Sauvegarde un DataFrame en Excel avec mise en forme professionnelle :
    en-tête bleu marine, alternance de couleurs, colonnes redimensionnées.
    """
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    with pd.ExcelWriter(str(path), engine="openpyxl") as writer:
        df.to_excel(writer, index=True, index_label="Ligne", sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        # Largeurs adaptées au contenu
        col_widths = {"A": 6, "B": 6, "C": 8, "D": 65, "E": 22, "F": 22}
        for col, w in col_widths.items():
            if col in [c.column_letter for c in ws[1]]:
                ws.column_dimensions[col].width = w

        # En-tête
        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        for cell in ws[1]:
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = border
        ws.row_dimensions[1].height = 30

        # Données
        amount_cols = {col.column for col in ws[1] if col.value in
                       ("AE (Milliers FCFA)", "CP (Milliers FCFA)", "AE", "CP")}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border    = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cell.column in amount_cols:
                    cell.number_format = "#,##0"
                    cell.alignment     = Alignment(horizontal="right", vertical="top")

        ws.freeze_panes = "A2"
